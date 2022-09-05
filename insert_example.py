import openpyxl

# Give the location of the file
path = "FORM_A_Tool_R.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# printing the value of first column
# This will find the question with id "insert_example" and replace the formulat of of that cell
# (which type is 'calculate') by 42. This value can then be used in notes or conditions further on in the form

for i in range(2, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=2)
    print(cell_obj.value)
    if (cell_obj.value == "insert_example"):
        cell_obj = sheet_obj.cell(row=i, column=12)
        print(cell_obj.value)
        cell_obj.value = "42"
        break
wb_obj.save("modified_copy.xlsx")
